# -*- coding: utf-8 -*-
"""
Excel解析模块 - 使用openpyxl（不依赖pandas）
"""

from openpyxl import load_workbook
from datetime import datetime


def parse_excel(file_path):
    """
    解析Excel文件，提取课程信息

    支持的列名：
    - 课程名称/课程名/课程/科目/name
    - 星期/周几/星期几/day
    - 开始时间/起始时间/上课时间/start_time
    - 结束时间/下课时间/end_time
    - 地点/教室/位置/location
    - 备注/说明/备注信息/remark
    """
    try:
        # 加载Excel文件
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 获取表头
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # 标准化列名映射
        column_mapping = {}
        for idx, col in enumerate(headers, 1):
            col_str = str(col).lower().strip()
            if any(keyword in col_str for keyword in ["课程", "科目", "name"]):
                column_mapping[idx] = "name"
            elif any(keyword in col_str for keyword in ["星期", "周几", "day"]):
                column_mapping[idx] = "day_of_week"
            elif any(keyword in col_str for keyword in ["开始", "起始", "start"]):
                column_mapping[idx] = "start_time"
            elif any(keyword in col_str for keyword in ["结束", "下课", "end"]):
                column_mapping[idx] = "end_time"
            elif any(
                keyword in col_str for keyword in ["地点", "教室", "位置", "location"]
            ):
                column_mapping[idx] = "location"
            elif any(keyword in col_str for keyword in ["备注", "说明", "remark"]):
                column_mapping[idx] = "remark"

        # 检查必需列
        required_fields = ["name", "day_of_week", "start_time", "end_time"]
        found_fields = [v for v in column_mapping.values()]
        missing_fields = [f for f in required_fields if f not in found_fields]

        if missing_fields:
            return {
                "success": False,
                "error": f"缺少必需列: {', '.join(missing_fields)}。请确保Excel包含：课程名称、星期、开始时间、结束时间",
            }

        courses = []
        errors = []

        # 解析每一行数据
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                row_data = {}
                for col_idx, value in enumerate(row, 1):
                    if col_idx in column_mapping:
                        row_data[column_mapping[col_idx]] = value

                course = parse_course_row(row_data)
                if course:
                    courses.append(course)
            except Exception as e:
                errors.append(f"第{row_idx}行解析失败: {str(e)}")

        wb.close()

        return {
            "success": True,
            "courses": courses,
            "count": len(courses),
            "errors": errors,
        }

    except Exception as e:
        return {"success": False, "error": f"Excel文件解析失败: {str(e)}"}


def parse_course_row(row_data):
    """解析单行课程数据"""
    # 课程名称
    name = str(row_data.get("name", "")).strip()
    if not name:
        raise ValueError("课程名称不能为空")

    # 星期处理
    day_of_week = parse_day_of_week(row_data.get("day_of_week"))
    if day_of_week is None:
        raise ValueError("星期格式不正确，请使用1-7或周一/星期一等格式")

    # 时间处理
    start_time = parse_time(row_data.get("start_time"))
    end_time = parse_time(row_data.get("end_time"))

    if not start_time:
        raise ValueError("开始时间格式不正确")
    if not end_time:
        raise ValueError("结束时间格式不正确")

    # 可选字段
    location = str(row_data.get("location", "") or "").strip()
    remark = str(row_data.get("remark", "") or "").strip()

    return {
        "name": name,
        "day_of_week": day_of_week,
        "start_time": start_time,
        "end_time": end_time,
        "location": location,
        "remark": remark,
    }


def parse_day_of_week(value):
    """解析星期"""
    if value is None:
        return None

    value_str = str(value).strip().lower()

    # 数字格式
    if value_str.isdigit():
        day = int(value_str)
        if 1 <= day <= 7:
            return day
        return None

    # 中文格式
    day_mapping = {
        "一": 1,
        "1": 1,
        "周一": 1,
        "星期一": 1,
        "mon": 1,
        "monday": 1,
        "二": 2,
        "2": 2,
        "周二": 2,
        "星期二": 2,
        "tue": 2,
        "tuesday": 2,
        "三": 3,
        "3": 3,
        "周三": 3,
        "星期三": 3,
        "wed": 3,
        "wednesday": 3,
        "四": 4,
        "4": 4,
        "周四": 4,
        "星期四": 4,
        "thu": 4,
        "thursday": 4,
        "五": 5,
        "5": 5,
        "周五": 5,
        "星期五": 5,
        "fri": 5,
        "friday": 5,
        "六": 6,
        "6": 6,
        "周六": 6,
        "星期六": 6,
        "sat": 6,
        "saturday": 6,
        "日": 7,
        "天": 7,
        "7": 7,
        "周日": 7,
        "星期天": 7,
        "星期日": 7,
        "sun": 7,
        "sunday": 7,
    }

    return day_mapping.get(value_str)


def parse_time(value):
    """解析时间"""
    if value is None:
        return None

    # 如果已经是datetime.time对象
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")

    value_str = str(value).strip()

    # 尝试多种时间格式
    time_formats = [
        "%H:%M",
        "%H:%M:%S",
        "%I:%M %p",
        "%I:%M:%S %p",
        "%H点%M分",
        "%H:%M:%S",
    ]

    for fmt in time_formats:
        try:
            dt = datetime.strptime(value_str, fmt)
            return dt.strftime("%H:%M")
        except ValueError:
            continue

    # 处理纯数字格式（如"800"表示"8:00"）
    if value_str.isdigit():
        if len(value_str) <= 2:
            # 只有小时
            return f"{int(value_str):02d}:00"
        elif len(value_str) == 3:
            # 如"830" -> "8:30"
            return f"{int(value_str[0]):01d}:{value_str[1:]}"
        elif len(value_str) == 4:
            # 如"1430" -> "14:30"
            return f"{value_str[:2]}:{value_str[2:]}"

    return None


def generate_template():
    """生成Excel模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "课程表"

    # 设置表头
    headers = ["课程名称", "星期", "开始时间", "结束时间", "地点", "备注"]
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["高等数学", "周一", "08:00", "09:40", "教学楼A101", "带教材"],
        ["大学英语", "2", "14:00", "15:40", "教学楼B203", "准备演讲"],
        ["计算机基础", "周三", "10:00", "11:40", "实验楼C305", "实验报告"],
    ]

    for example in examples:
        ws.append(example)

    # 设置列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    # 添加边框
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(
        min_row=1, max_row=len(examples) + 1, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = thin_border

    return wb
